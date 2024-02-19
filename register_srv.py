from pathlib import PurePath
from typing import cast
import openpyxl
from models import Register, RegisterRow
from openpyxl.worksheet.worksheet import Worksheet


def load_from_file(path: str) -> Register:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheet = cast(Worksheet, wb.active)
    data = [
        RegisterRow(str(row[0]), row[2], row[4], row[5])
        for row in sheet.iter_rows(values_only=True, min_col=2, min_row=3)
        if row[0]
    ]
    register = Register()
    register.data = data
    return register


def save_to_file(register: Register, path: str):
    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    if ws is None:
        raise ValueError("Active sheet not found")
    ws.title = PurePath(path).stem
    for row in register:
        ws.append(row)
    wb.save(path)
